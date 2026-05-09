import os
import re

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC_DIR = os.path.join("teenieping", "img_assets", "teenieping_imgs")
MD_OUT = "teenieping_list.md"
XLSX_OUT = "teenieping_list.xlsx"

PAT = re.compile(r"^(TP-\d{3})_(.+)\.webp$")

# 시즌 표 (전체 일람용)
SEASONS_OVERVIEW = [
    ("1기", "큐브 티니핑"),
    ("2기", "보석 티니핑"),
    ("3기", "열쇠 티니핑"),
    ("4기", "디저트 티니핑"),
    ("5기", "스타 티니핑"),
    ("6기", "프린세스&프린스 티니핑"),
    ("-", "특별 티니핑"),
]

# 추가 캐릭터 (원본 webp 폴더에 파일이 없는 항목)
EXTRA_ROWS = [
    ("TP-151", "로미핑"),
]

# namu.wiki 4.1.x 절 구조 기반 정본 매핑
# 캐릭터 이름 -> (기수, 등급)
ROYAL = {
    "1기": ["하츄핑", "바로핑", "아자핑", "차차핑", "라라핑", "해핑"],
    "2기": ["조아핑", "방글핑", "믿어핑"],
    "3기": ["꾸래핑", "솔찌핑", "나나핑"],
    "4기": ["포실핑", "말랑핑", "샤샤핑"],
    "5기": ["빛나핑", "초롱핑", "빤짝핑", "왕자핑"],
    "6기": ["사뿐핑", "아름핑", "뽀니핑"],
}
LEGEND = {
    "3기": ["행운핑"],
    "4기": ["새콤핑", "달콤핑"],
    "5기": ["오로라핑"],
    "6기": ["이클립스핑", "다이아나핑"],
}

NORMAL = {
    "1기": [
        "키키핑", "아잉핑", "부끄핑", "부투핑", "깜빡핑", "띠용핑", "주르핑", "차나핑",
        "따라핑", "나르핑", "무셔핑", "투투핑", "차캐핑", "떠벌핑", "다조핑", "화나핑",
        "시러핑", "바네핑", "악동핑", "덜덜핑", "그림핑", "무거핑", "꺼꿀핑", "씽씽핑",
        "베베핑", "코자핑", "딱풀핑", "모야핑", "토이핑", "또까핑", "플라핑",
        "노라핑&노리핑", "아휴핑", "똑똑핑", "꽁꽁핑", "찌릿핑", "홀로핑", "앙대핑",
    ],
    "2기": [
        "까르핑", "아야핑", "소원핑", "토닥핑", "쪼꼼핑", "싹싹핑", "맛나핑", "포근핑",
        "메모핑", "다해핑", "공쥬핑", "짝짝핑", "주네핑", "뚝딱핑", "발레핑", "원더핑",
    ],
    "3기": [
        "빨리핑", "얌얌핑", "뜨거핑", "삐뽀핑", "간호핑", "힘내핑", "고쳐핑", "아라핑",
        "패션핑", "꼼딱핑", "퐁당핑", "파티핑", "꾸며핑", "삐짐핑", "아아핑", "빙글핑",
    ],
    "4기": [
        "뿌뿌핑", "캔디핑", "머랭핑", "샌드핑", "또너핑", "와플핑", "롤리핑", "마카핑",
        "핫케핑", "커핑&머핑", "요거핑", "눈꽃핑", "푸딩핑", "멜로핑", "쪼꼬핑",
    ],
    "5기": [
        "깡총핑", "훌라핑", "나눔핑", "딩동핑", "나그네핑", "고고핑", "뿌쵸핑", "고마핑",
        "아롱핑&다롱핑", "아롱핑", "다롱핑", "뽀송핑", "유리핑", "함께핑", "댄스핑",
        "여우핑", "루루핑", "몰래핑", "뽀뽀핑",
    ],
    "6기": [
        "뽀득핑", "차밍핑", "나비핑", "실크핑", "스노우핑", "이슬핑", "쿨쿨핑", "슈슈핑",
        "롱롱핑", "큐핑", "야옹핑", "샤를핑", "트롯핑", "깨굴핑", "젠틀핑",
    ],
}

# 빠른 조회용 역인덱스 (이름 -> (기수, 등급))
NAME_INDEX: dict[str, tuple[str, str]] = {}
for gen, names in ROYAL.items():
    for n in names:
        NAME_INDEX[n] = (gen, "로열")
for gen, names in LEGEND.items():
    for n in names:
        NAME_INDEX[n] = (gen, "레전드")
for gen, names in NORMAL.items():
    for n in names:
        NAME_INDEX[n] = (gen, "일반")

# 일련번호별 특수 케이스 (override + 비고)
SPECIAL = {
    "TP-064": {
        "기수": "2기",
        "종류": "보석 티니핑",
        "등급": "일반",
        "비고": "다해핑(TP-065)이 가면을 쓴 상태 (동일 캐릭터). namu.wiki 카운트에서는 별도 종으로 집계하지 않음.",
    },
    "TP-130": {
        "기수": "-",
        "종류": "특별 티니핑",
        "등급": "특별",
        "비고": "사랑의 하츄핑 (극장판 1기, 2024). 본편 시리즈에는 미등장.",
    },
    "TP-151": {
        "기수": "-",
        "종류": "특별 티니핑",
        "등급": "특별",
        "비고": "로미가 티니핑으로 변한 모습.",
    },
    "TP-120": {
        "기수": "5기",
        "종류": "스타 티니핑",
        "등급": "일반",
        "비고": "다롱핑(TP-121)과 쌍둥이. namu.wiki에서는 「아롱핑&다롱핑」으로 합쳐서 표기.",
    },
    "TP-121": {
        "기수": "5기",
        "종류": "스타 티니핑",
        "등급": "일반",
        "비고": "아롱핑(TP-120)과 쌍둥이. namu.wiki에서는 「아롱핑&다롱핑」으로 합쳐서 표기.",
    },
    "TP-037": {
        "기수": "1기",
        "종류": "큐브 티니핑",
        "등급": "일반",
        "비고": "노라핑·노리핑 2종이 한 항목으로 묶임 (원본 webp 파일 통합).",
    },
    "TP-099": {
        "기수": "4기",
        "종류": "디저트 티니핑",
        "등급": "일반",
        "비고": "커핑·머핑 2종이 한 항목으로 묶임 (원본 webp 파일 통합).",
    },
}

KIND_BY_GEN = {
    "1기": "큐브 티니핑",
    "2기": "보석 티니핑",
    "3기": "열쇠 티니핑",
    "4기": "디저트 티니핑",
    "5기": "스타 티니핑",
    "6기": "프린세스&프린스 티니핑",
    "극장판": "특별 티니핑",
}

def classify(sn: str, name: str):
    if sn in SPECIAL:
        s = SPECIAL[sn]
        return s["기수"], s["종류"], s["등급"], s["비고"]
    if name in NAME_INDEX:
        gen, grade = NAME_INDEX[name]
        kind = KIND_BY_GEN[gen]
        if grade == "로열":
            note = f"{gen} 로열 티니핑"
        elif grade == "레전드":
            note = f"{gen} 레전드 티니핑"
        else:
            note = ""
        return gen, kind, grade, note
    return "?", "?", "?", "namu.wiki 정본 매핑에 없음"

def collect():
    rows = []
    for fn in os.listdir(SRC_DIR):
        m = PAT.match(fn)
        if not m:
            continue
        sn, name = m.group(1), m.group(2)
        gen, kind, grade, note = classify(sn, name)
        rows.append((sn, name, gen, kind, grade, note))
    for sn, name in EXTRA_ROWS:
        gen, kind, grade, note = classify(sn, name)
        rows.append((sn, name, gen, kind, grade, note))
    rows.sort(key=lambda x: x[0])
    return rows

def write_md(rows):
    lines = []
    lines.append("# 티니핑 캐릭터 목록")
    lines.append("")
    lines.append(f"총 {len(rows)}종 (TP-001 ~ TP-150 중 TP-038, TP-100 결번 / TP-151 로미핑은 특별 추가)")
    lines.append("")
    lines.append("출처: namu.wiki 「티니핑」 문서 4.1절 (시즌별 분류) 기준. 복수 시즌 등장 캐릭터는 처음 등장한 기수로 분류.")
    lines.append("")
    lines.append("## 기수 구분")
    lines.append("")
    lines.append("| 기수 | 종류 |")
    lines.append("|---|---|")
    for gen, kind in SEASONS_OVERVIEW:
        lines.append(f"| {gen} | {kind} |")
    lines.append("")
    lines.append("## 등급 표기")
    lines.append("")
    lines.append("- **로열**: 각 기수의 주역 티니핑 (왕족/주연 라인)")
    lines.append("- **레전드**: 3기부터 도입된 등급. 4기는 새콤핑·달콤핑, 5기는 오로라핑 등")
    lines.append("- **일반**: 그 외 일반 티니핑")
    lines.append("- **특별**: 본편 시리즈 외(극장판 한정 등) 등장 캐릭터")
    lines.append("")
    lines.append("## 전체 목록")
    lines.append("")
    lines.append("| 일련번호 | 이름 | 기수 | 종류 | 등급 | 비고 |")
    lines.append("|---|---|---|---|---|---|")
    for sn, name, gen, kind, grade, note in rows:
        lines.append(f"| {sn} | {name} | {gen} | {kind} | {grade} | {note} |")
    lines.append("")
    with open(MD_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

def write_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "티니핑 목록"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6B5B95")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["일련번호", "이름", "기수", "종류", "등급", "비고"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for i, (sn, name, gen, kind, grade, note) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=sn).alignment = center
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=gen).alignment = center
        ws.cell(row=i, column=4, value=kind)
        ws.cell(row=i, column=5, value=grade).alignment = center
        ws.cell(row=i, column=6, value=note).alignment = left

    widths = [12, 20, 8, 24, 10, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.auto_filter.ref = f"A1:F{len(rows)+1}"
    ws.freeze_panes = "A2"

    wb.save(XLSX_OUT)

def main():
    rows = collect()
    print(f"collected {len(rows)} entries")
    # check unmapped
    miss = [(r[0], r[1]) for r in rows if r[2] == "?"]
    if miss:
        print(f"WARN: {len(miss)} unmapped:")
        for sn, n in miss:
            print(f"  {sn} {n}")
    # season distribution
    from collections import Counter
    c = Counter(r[2] for r in rows)
    for gen, kind in SEASONS_OVERVIEW:
        print(f"  {gen} {kind}: {c.get(gen, 0)}")
    # grade distribution
    g = Counter(r[4] for r in rows)
    print(f"  로열/레전드/일반/극장판: {g}")
    write_md(rows)
    write_xlsx(rows)
    print(f"wrote {MD_OUT} and {XLSX_OUT}")

if __name__ == "__main__":
    main()
