"""
카봇 카탈로그 빌더 — 엑셀(.xlsx) + 시각화(.html) 동시 생성.
한 번만 쓰는 스크립트. carbot/ 디렉터리에서 실행.
"""
import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
SRC  = BASE / "img_carbot_src"

# ─────────────────────────────────────────────────────────────────────────
# 1. 데이터 정의
# ─────────────────────────────────────────────────────────────────────────

ROLE_TABLE = [
    ("standalone", "개별", "합체와 무관, 단독 활약",
     "카봇 호크, 카봇 빅터, 카봇 라란자, 카봇 팔로, 카봇 베어하이더, 알카봇, 카봇 트루"),
    ("member", "멤버", "합체의 구성요소 + 단독 모습",
     "카봇 에이스, 카봇 스카이, 카봇 가드, 카봇 자크, 카봇 마이스터"),
    ("combined", "합체", "멤버들이 합쳐진 거대 로봇",
     "카봇 펜타스톰, 카봇 마이티가드, 카봇 로드세이버, 잭슈트뱅"),
]

MODE_TABLE = [
    ("robot",         "인간형/휴머노이드 로봇",              "거의 모든 카봇"),
    ("vehicle",       "자동차/항공기/선박 등 탈것",          "카봇 스카이의 해치백, 카봇 에이스의 SUV"),
    ("beast",         "동물 형태 (전신 또는 하체)",          "카봇 팔로(하체 물소), 베어하이더(곰), 디고베어"),
    ("animal_full",   "완전한 진짜 동물 모드 (트리플 체인저)", "카봇 팔로 (완전 물소 모드)"),
    ("power",         "합체 로봇의 분노형/강화형",            "로어세이버 파워 모드 (= 로드세이버 곰형)"),
    ("armed",         "특정 무장 장착 형태 (변형 아님)",      "펜타스톰 X 빅큐브 부스터/기어 모드"),
    ("combined_only", "단독 변형 없이 합체 시에만 변형",       "마이스터/아티 (작중 합체용 변형만)"),
]

VARIATION_AXIS_TABLE = [
    ("x",            "시즌 진화",       "스카이 → 스카이 S.W.A.T X (시즌 9)",            "디자인 강화, 새 비클, 새 기술"),
    ("crystal",      "컬러/소재",       "크리스탈 카봇 스카이 S.W.A.T X (시즌 13)",      "반투명 재질, 빨간 포인트"),
    ("gold",         "컬러/소재",       "카봇 펜타스톰 황금특공대 (시즌 10)",            "황금 맥기, 블랙+골드 투톤"),
    ("allstar",      "컬러/소재",       "카봇 로드세이버 올스타 (시즌 15)",              "반투명 어레인지"),
    ("refine",       "컬러/소재",       "카봇 펜타스톰 리파인 (시즌 16)",                "페이스리프트 디자인"),
    ("bigcube",      "무장 추가",       "카봇 펜타스톰 X 빅큐브 (시즌 14)",              "빅큐브 무장 장착"),
    ("lifecannon",   "무장 추가",       "카봇 펜타스톰 X 라이프 캐논 모드 (시즌 14)",    "새니타이저 캐논 + 콰트로 클로"),
    ("power_mode",   "합체 방식 변경",  "로드세이버 → 로어세이버 파워 모드",             "같은 멤버, 다른 합체 결과 (인간형↔곰형)"),
    ("plus_member",  "멤버 추가/교체",  "펜타스톰 X (5단) → X 라이프 캐논 모드 (6단)",   "멤버 1개 추가로 그레이트 합체"),
    ("cross",        "크로스 콤비",     "마하피스/브레이피스/마하로드/브레이로드",        "상체-하체 멤버 교차로 4가지 합체"),
    ("vehicle_var",  "비클 변종",       "스카이(기본)/스카이 S.W.A.T/댄디(구급차)",      "같은 멤버의 다른 비클 모드"),
]

# 나무위키 페이지에서 직접 확인한 변종 — (로봇명, 변종ID, 변종한국명, 시즌범위, 모드들, 비고)
VERIFIED_VARIATIONS = [
    # 카봇 스카이 (member + standalone)
    ("카봇 스카이", "basic",            "카봇 스카이",                       "1~16",  "robot, vehicle",                "해치백 (현대 벨로스터 1세대), 주황색"),
    ("카봇 스카이", "swat",             "카봇 스카이 S.W.A.T",               "1~",    "robot, vehicle",                "해치백 SWAT 경찰차, 검정"),
    ("카봇 스카이", "swat_x",           "카봇 스카이 S.W.A.T X",             "9~15",  "robot, vehicle",                "SWAT 기동차량 (가상), 검정"),
    ("카봇 스카이", "swat_x_crystal",   "크리스탈 카봇 스카이 S.W.A.T X",    "13",    "robot",                          "반투명 재질, 빨간 포인트"),

    # 카봇 팔로 (standalone, 극장판)
    ("카봇 팔로",   "basic",            "카봇 팔로",                         "극장판",  "robot, beast, animal_full",     "아프리카물소, 트리플 체인저 (역대 유일)"),

    # 카봇 로드세이버 (combined)
    ("카봇 로드세이버", "basic",        "카봇 로드세이버",                   "2~",     "robot, power",                  "인간형 ↔ 로어세이버(곰형). 멤버: 세이버, 마이스터 택시, 아티 택시"),
    ("카봇 로드세이버", "allstar",      "카봇 로드세이버 올스타",            "15~",    "robot, power",                  "반투명 어레인지. 새 뱅크씬"),

    # 카봇 펜타스톰 (combined)
    ("카봇 펜타스톰", "basic",          "카봇 펜타스톰",                     "1~3,16", "robot",                          "5단. 멤버: 스톰/에이스 레스큐/프론 경찰차/댄디 구급차/스카이 SWAT"),
    ("카봇 펜타스톰", "gold",           "카봇 펜타스톰 황금특공대",          "10~11",  "robot",                          "기본과 동일 5종, 컬러만 황금/블랙"),
    ("카봇 펜타스톰", "refine",         "카봇 펜타스톰 리파인",              "16",     "robot",                          "페이스리프트 디자인 (라이선스 회피)"),

    # 카봇 펜타스톰 X (combined)
    ("카봇 펜타스톰 X", "basic",        "카봇 펜타스톰 X",                   "9~15",   "robot",                          "X 강화판 5단. 멤버: 스톰 X/에이스 레스큐 X/프론 폴리스 X/댄디 앰뷸런스 X/스카이 SWAT X"),
    ("카봇 펜타스톰 X", "crystal",      "크리스탈 카봇 펜타스톰 X",          "13",     "robot",                          "13기 한정. 호크 X와 함께 크리스탈 업그레이드"),
    ("카봇 펜타스톰 X", "bigcube",      "카봇 펜타스톰 X 빅큐브",            "14~",    "robot, armed",                  "빅큐브 무장 (부스터 모드/기어 모드)"),
    ("카봇 펜타스톰 X", "lifecannon",   "카봇 펜타스톰 X 라이프 캐논 모드",  "14~",    "robot",                          "6단 그레이트 합체. 멤버 +라이프 X. 새니타이저 캐논, 콰트로 클로"),
]

# build_carbot_images.py 의 COMBO_COMPONENTS 미러 (실측 합체-멤버 그래프)
COMBO_COMPONENTS = {
    # 시즌 1
    "카봇 펜타스톰":     ["에이스", "프론", "댄디", "스카이", "스톰"],
    "카봇 펜타스톰 X":   ["라이프 X"],
    # 시즌 2
    "카봇 로드세이버":   ["아티", "마이스터", "세이버"],
    "카봇 삼총사":       ["나이트", "루크", "폰"],
    # 시즌 3
    "카봇 마이티가드":   ["레디", "큐어", "가드", "헬프"],
    "카봇 K-캅스":       ["썬더", "스피드", "터보", "캅스"],
    # 시즌 4
    "카봇 슈퍼패트론":   ["패트론", "스키드", "다이어", "리프"],
    "카봇 패트론S":      ["패트론", "스키드"],
    "카봇 다이어EX":     ["다이어", "리프"],
    # 시즌 5
    "카봇 프라우드제트": ["프라우드", "제스티"],
    "카봇 스타블래스터": ["블래스터", "스타비"],
    "카봇 아머다이저":   ["킹다이저", "아머포스"],
    # 시즌 6
    "카봇 럭키펀치":     ["럭키", "펀치"],
    # 시즌 10 가디언트 / 자이언트 로더
    "카봇 가디언트":     ["덤피", "브로피"],
    "카봇 자이언트 로더":["페이저", "소닉붐", "로더"],
    # 시즌 10 뱅 시리즈
    "잭슈트뱅":          ["잭슈트", "타우", "퓨처버스"],
    "페로뱅":            ["페로", "파우", "퓨처버스"],
    "나노티스뱅":        ["나노티스", "도우", "퓨처버스"],
    "썬런뱅":            ["썬런", "미우", "파이트라"],
    "호스퍼스뱅":        ["호스퍼스", "바우", "파이트라"],
    "자크뱅":            ["자크", "바우", "파이트라"],
    "자크레인뱅":        ["자크", "자크레인"],
    # 시즌 11~17
    "카봇 킹가이더":     ["핫가이더", "쿨가이더"],
    "카봇 하이퍼캅스":   ["드로캅", "이그리붐", "푸르디붐", "서포티붐"],
    "카봇 사파리세이버": ["드럼킹", "호크하이", "하울러", "타이거붐", "스터번"],
    "카봇 스타가디언":   ["폴리스타", "스카이스타", "아이언스타", "파이어스타"],
    "그랜드 카봇 X":     ["카봇 X", "엑스 파이터", "엑스 트레인"],
    "카봇 패트롤가이즈": ["히트가이", "코드가이"],
    "카봇 스카이 트리오":["제트밴더", "에어밴더", "헬리밴더"],
    # 크로스 콤비
    "카봇 마하피스":     ["마하", "피스"],
    "카봇 브레이피스":   ["브레이브", "피스"],
    "카봇 마하로드":     ["마하", "로드"],
    "카봇 브레이로드":   ["브레이브", "로드"],
}

FILENAME_RULE_EXAMPLES = [
    ("카봇 스카이__basic__robot.webp",            "기본 스카이 로봇 모드"),
    ("카봇 스카이__basic__vehicle.webp",          "기본 스카이 비클 모드 (벨로스터)"),
    ("카봇 스카이__swat__robot.webp",             "SWAT 변종 로봇"),
    ("카봇 스카이__swat__vehicle.webp",           "SWAT 변종 경찰차"),
    ("카봇 스카이__swat_x__robot.webp",           "X 강화판 로봇"),
    ("카봇 스카이__swat_x__vehicle.webp",         "X 강화판 비클"),
    ("카봇 스카이__swat_x_crystal__robot.webp",   "크리스탈 변종"),
    ("카봇 팔로__basic__robot.webp",              "팔로 로봇 모드"),
    ("카봇 팔로__basic__beast.webp",              "팔로 비스트(하체 물소)"),
    ("카봇 팔로__basic__animal_full.webp",        "팔로 완전 물소 (트리플 체인저)"),
    ("카봇 로드세이버__basic__robot.webp",        "로드세이버 인간형"),
    ("카봇 로드세이버__basic__power.webp",        "로어세이버 (곰형)"),
    ("카봇 로드세이버__allstar__robot.webp",      "올스타 인간형"),
    ("카봇 로드세이버__allstar__power.webp",      "올스타 곰형"),
    ("카봇 펜타스톰__basic__robot.webp",          "기본 펜타스톰"),
    ("카봇 펜타스톰__gold__robot.webp",           "황금특공대"),
    ("카봇 펜타스톰__refine__robot.webp",         "리파인"),
    ("카봇 펜타스톰 X__basic__robot.webp",        "펜타스톰 X 기본"),
    ("카봇 펜타스톰 X__crystal__robot.webp",      "크리스탈 펜타스톰 X"),
    ("카봇 펜타스톰 X__bigcube__robot.webp",      "빅큐브 장착"),
    ("카봇 펜타스톰 X__lifecannon__robot.webp",   "라이프 캐논 (6단)"),
    ("카봇 마하피스__cross__robot.webp",          "마하+피스 교차 합체"),
]


# ─────────────────────────────────────────────────────────────────────────
# 2. 현재 img_carbot_src 상태 스캔
# ─────────────────────────────────────────────────────────────────────────

def scan_inventory():
    all_files = [f for f in SRC.iterdir() if f.is_file() and f.suffix == '.webp']
    h_to_files = defaultdict(list)
    for f in all_files:
        h = hashlib.md5(f.read_bytes()).hexdigest()
        h_to_files[h].append(f.name)

    rows = []
    for f in sorted(all_files):
        h = hashlib.md5(f.read_bytes()).hexdigest()
        if f.stem.endswith('_zz'):
            suffix = '_zz'
        elif f.stem.endswith('_xxyy'):
            suffix = '_xxyy'
        elif f.stem.endswith('_xx'):
            suffix = '_xx'
        else:
            suffix = '(clean)'
        sib = [n for n in h_to_files[h] if n != f.name]
        rows.append((f.name, suffix, f.stat().st_size // 1024, h[:8], len(sib), ', '.join(sib[:3]) + ('...' if len(sib) > 3 else '')))

    # 카테고리 카운트
    cats = {'_zz': 0, '_xxyy': 0, '_xx': 0, '(clean)': 0}
    for r in rows:
        cats[r[1]] += 1

    # 중복 그룹
    dups = {h: names for h, names in h_to_files.items() if len(names) > 1}

    return rows, cats, dups


# ─────────────────────────────────────────────────────────────────────────
# 3. Excel 빌드
# ─────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF'))


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(out_path):
    wb = Workbook()
    wb.remove(wb.active)

    # 01 역할분류
    ws = wb.create_sheet("01_역할분류")
    ws.append(["키(role)", "이름", "정의", "예시"])
    for row in ROLE_TABLE:
        ws.append(list(row))
    style_header(ws, 4)
    autofit(ws, [14, 10, 38, 70])

    # 02 모드분류
    ws = wb.create_sheet("02_모드분류")
    ws.append(["모드 키", "설명", "예시"])
    for row in MODE_TABLE:
        ws.append(list(row))
    style_header(ws, 3)
    autofit(ws, [16, 40, 50])

    # 03 변종축
    ws = wb.create_sheet("03_변종축")
    ws.append(["변종 키", "변종 축 분류", "예시", "변경 내용"])
    for row in VARIATION_AXIS_TABLE:
        ws.append(list(row))
    style_header(ws, 4)
    autofit(ws, [16, 16, 50, 38])

    # 04 검증된변종
    ws = wb.create_sheet("04_검증된변종")
    ws.append(["로봇명", "변종 ID", "변종 한국명", "시즌 범위", "지원 모드", "비고"])
    for row in VERIFIED_VARIATIONS:
        ws.append(list(row))
    style_header(ws, 6)
    autofit(ws, [18, 22, 32, 12, 28, 70])

    # 05 합체관계
    ws = wb.create_sheet("05_합체관계")
    ws.append(["합체 로봇", "구성 멤버 수", "구성 멤버"])
    for combo, members in COMBO_COMPONENTS.items():
        ws.append([combo, len(members), ", ".join(members)])
    style_header(ws, 3)
    autofit(ws, [22, 12, 60])

    # 06 파일현황
    rows, cats, dups = scan_inventory()
    ws = wb.create_sheet("06_파일현황")
    ws.append(["파일명", "접미사", "크기(KB)", "MD5", "중복수", "동일 이미지 파일"])
    for row in rows:
        ws.append(list(row))
    style_header(ws, 6)
    autofit(ws, [40, 10, 10, 12, 8, 50])

    # 07 인벤토리요약
    ws = wb.create_sheet("07_인벤토리요약")
    ws.append(["접미사", "개수", "의미"])
    cat_meaning = {
        '_zz':    '사용자 수동 수정 (고유 이미지)',
        '_xxyy':  '자동 마킹: 합체 로봇 (중복)',
        '_xx':    '자동 마킹: 구성원 로봇 (중복)',
        '(clean)':'단독 이미지 (정상)',
    }
    for k in ['(clean)', '_zz', '_xxyy', '_xx']:
        ws.append([k, cats[k], cat_meaning[k]])
    ws.append([])
    ws.append(["중복 그룹 수", len(dups), "MD5 기준 2개 이상 동일 그룹"])
    ws.append(["중복 파일 합계", sum(len(n) for n in dups.values()), ""])
    ws.append(["총 파일 수", sum(cats.values()), ""])
    style_header(ws, 3)
    autofit(ws, [14, 10, 50])

    # 08 중복그룹
    ws = wb.create_sheet("08_중복그룹")
    ws.append(["그룹 크기", "MD5 (앞 8자)", "파일들"])
    for h, names in sorted(dups.items(), key=lambda x: -len(x[1])):
        ws.append([len(names), h[:8], ", ".join(sorted(names))])
    style_header(ws, 3)
    autofit(ws, [12, 14, 100])

    # 09 파일명규칙
    ws = wb.create_sheet("09_파일명규칙")
    ws.append(["예시 파일명", "의미"])
    for row in FILENAME_RULE_EXAMPLES:
        ws.append(list(row))
    style_header(ws, 2)
    autofit(ws, [50, 50])

    # 00 표지 (목차)
    cover = wb.create_sheet("00_표지", 0)
    cover.append(["카봇 카탈로그 (Carbot Catalog)"])
    cover['A1'].font = Font(bold=True, size=18, color="305496")
    cover.append([f"생성일: 2026-05-18"])
    cover.append([])
    cover.append(["시트 목차"])
    cover['A4'].font = Font(bold=True, size=14)
    toc = [
        ("01_역할분류",     "역할 3분류 (개별 / 멤버 / 합체)"),
        ("02_모드분류",     "모드 7종 (robot / vehicle / beast / animal_full / power / armed / combined_only)"),
        ("03_변종축",       "변종 축 11종 (x / crystal / gold / refine / bigcube / lifecannon ...)"),
        ("04_검증된변종",   "나무위키 확인 완료 변종 13개 (스카이/팔로/로드세이버/펜타스톰 패밀리)"),
        ("05_합체관계",     "합체 로봇 33종 → 구성 멤버 그래프"),
        ("06_파일현황",     "img_carbot_src/ 113개 파일 현재 상태"),
        ("07_인벤토리요약", "접미사별 집계"),
        ("08_중복그룹",     "MD5 기준 중복 그룹"),
        ("09_파일명규칙",   "새 파일명 규칙 (<이름>__<변종>__<모드>.webp) 예시"),
    ]
    for name, desc in toc:
        cover.append([name, desc])
    style_header(cover, 0)  # header 스킵
    autofit(cover, [22, 80])

    wb.save(out_path)
    print(f"[xlsx] {out_path}")


# ─────────────────────────────────────────────────────────────────────────
# 4. HTML 시각화 빌드 (Plotly.js CDN)
# ─────────────────────────────────────────────────────────────────────────

def build_html(out_path, cats, dups):
    # 데이터 준비
    inventory_data = [
        ('단독 정상 (clean)',         cats['(clean)'], '#5cb85c'),
        ('합체 자동 마킹 (_xxyy)',     cats['_xxyy'],   '#f0ad4e'),
        ('구성원 자동 마킹 (_xx)',     cats['_xx'],     '#d9534f'),
        ('수동 수정 완료 (_zz)',       cats['_zz'],     '#5bc0de'),
    ]

    # 중복 그룹 분포
    dup_sizes = sorted([len(n) for n in dups.values()], reverse=True)
    dup_size_distribution = defaultdict(int)
    for s in dup_sizes:
        dup_size_distribution[s] += 1

    # 펜타스톰 패밀리 sunburst 데이터
    penta_sunburst_data = {
        'labels': [
            '펜타스톰 패밀리',
            '기본 펜타스톰', '펜타스톰 X',
            '기본', '황금특공대', '리파인',
            '펜타스톰 X 기본', '크리스탈', '빅큐브', '라이프 캐논 (6단)',
        ],
        'parents': [
            '',
            '펜타스톰 패밀리', '펜타스톰 패밀리',
            '기본 펜타스톰', '기본 펜타스톰', '기본 펜타스톰',
            '펜타스톰 X', '펜타스톰 X', '펜타스톰 X', '펜타스톰 X',
        ],
        'values': [0, 3, 4, 1, 1, 1, 1, 1, 1, 1],
    }

    # 합체-멤버 Sankey 데이터
    combos = list(COMBO_COMPONENTS.keys())
    members_seen = []
    sankey_source, sankey_target, sankey_value = [], [], []
    label_to_idx = {}
    labels = []

    def get_idx(name):
        if name not in label_to_idx:
            label_to_idx[name] = len(labels)
            labels.append(name)
        return label_to_idx[name]

    # 합체로봇 노드 먼저 (왼쪽)
    for c in combos:
        get_idx(c)
    # 멤버 노드 (오른쪽)
    for c, members in COMBO_COMPONENTS.items():
        for m in members:
            mn = m if (m.startswith('카봇 ') or m.startswith('그랜드 카봇 ')) else f'카봇 {m}'
            mi = get_idx(mn)
            ci = get_idx(c)
            sankey_source.append(ci)
            sankey_target.append(mi)
            sankey_value.append(1)

    # 색상 — 합체는 짙은 파랑, 멤버는 회색
    combo_set = set(combos)
    node_colors = ['#1f77b4' if l in combo_set else '#9ca3af' for l in labels]

    # 역할 분포 (estimated)
    role_estimate = {
        '합체 로봇 (조회된 33종)': 33,
        '구성원 (자동 수집 60종)': 60,
        '단독/극장판/올스타 등': 25,
    }

    # 검증 진행률
    verified_count = len(set(v[0] for v in VERIFIED_VARIATIONS))
    total_estimate = 118
    verify_progress = {
        '검증 완료': verified_count,
        '미검증': total_estimate - verified_count,
    }

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="apple-mobile-web-app-capable" content="yes">
<title>카봇 카탈로그 시각화 — 2026-05-18</title>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
<style>
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Apple SD Gothic Neo', 'Noto Sans KR', sans-serif;
    margin: 0; padding: 20px; background: #f5f7fa; color: #2c3e50;
  }}
  h1 {{ color: #305496; margin-bottom: 4px; }}
  .subtitle {{ color: #6c757d; margin-bottom: 24px; }}
  .grid {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 20px;
  }}
  .card {{
    background: white; border-radius: 12px; padding: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }}
  .card.full {{ grid-column: 1 / -1; }}
  .card h2 {{ margin-top: 0; color: #305496; font-size: 18px; }}
  .card .desc {{ color: #6c757d; font-size: 13px; margin-bottom: 12px; }}
  .stats {{
    display: flex; gap: 16px; margin-bottom: 16px;
  }}
  .stat {{
    flex: 1; background: #e8eef7; padding: 12px; border-radius: 8px; text-align: center;
  }}
  .stat .num {{ font-size: 28px; font-weight: bold; color: #305496; }}
  .stat .label {{ font-size: 12px; color: #6c757d; margin-top: 4px; }}
  .legend-note {{
    font-size: 12px; color: #6c757d; margin-top: 8px;
    padding: 8px 12px; background: #f8f9fa; border-left: 3px solid #305496; border-radius: 4px;
  }}
</style>
</head>
<body>

<h1>카봇 카탈로그 시각화</h1>
<div class="subtitle">img_carbot_src/ 인벤토리 + 나무위키 검증 결과 + 합체 관계 그래프</div>

<div class="stats">
  <div class="stat"><div class="num">{sum(cats.values())}</div><div class="label">총 이미지 파일</div></div>
  <div class="stat"><div class="num">{cats['(clean)']}</div><div class="label">단독 정상</div></div>
  <div class="stat"><div class="num">{cats['_zz']}</div><div class="label">_zz 수동 수정</div></div>
  <div class="stat"><div class="num">{cats['_xxyy'] + cats['_xx']}</div><div class="label">중복 미해결</div></div>
  <div class="stat"><div class="num">{len(dups)}</div><div class="label">중복 그룹</div></div>
  <div class="stat"><div class="num">{verified_count}</div><div class="label">검증된 로봇</div></div>
</div>

<div class="grid">

  <div class="card">
    <h2>1. img_carbot_src 인벤토리 상태</h2>
    <div class="desc">접미사별 파일 분포 — 무엇이 정상이고 무엇이 손봐야 하는지</div>
    <div id="chart_inventory" style="height:340px"></div>
  </div>

  <div class="card">
    <h2>2. 검증 진행률</h2>
    <div class="desc">나무위키 페이지로 변종 확인을 완료한 로봇 / 미검증 로봇</div>
    <div id="chart_verify" style="height:340px"></div>
  </div>

  <div class="card">
    <h2>3. 중복 그룹 분포</h2>
    <div class="desc">한 그룹에 동일 이미지가 몇 개씩 묶여 있는지 (그룹 크기별 개수)</div>
    <div id="chart_dupsize" style="height:340px"></div>
  </div>

  <div class="card">
    <h2>4. 펜타스톰 패밀리 트리</h2>
    <div class="desc">사용자가 _zz로 정리한 펜타스톰 라인업의 계층 — 검증 완료 7개 변종</div>
    <div id="chart_penta" style="height:340px"></div>
  </div>

  <div class="card full">
    <h2>5. 합체-멤버 관계 (Sankey)</h2>
    <div class="desc">왼쪽: 합체 로봇 33종. 오른쪽: 구성원 로봇. 한 멤버가 여러 합체에 들어가는 경우도 한눈에.</div>
    <div id="chart_sankey" style="height:880px"></div>
  </div>

  <div class="card full">
    <h2>6. 카봇 스카이 변종-모드 매트릭스 (검증 완료 예시)</h2>
    <div class="desc">한 멤버 로봇이 가질 수 있는 (변종 × 모드) 조합. 녹색=확인된 이미지 존재, 회색=페이지에 없음.</div>
    <div id="chart_skymatrix" style="height:300px"></div>
    <div class="legend-note">
      <b>핵심 관찰</b>: 한 페이지에 비클/로봇 이미지가 최소 6장 (변종 4 × 모드 2 - 일부). 현재 크롤러는 og:image 1장만 가져옴 → 5장 누락.
    </div>
  </div>

</div>

<script>
// 1. 인벤토리 도넛
Plotly.newPlot('chart_inventory', [{{
  type: 'pie', hole: 0.55,
  labels: {json.dumps([d[0] for d in inventory_data], ensure_ascii=False)},
  values: {json.dumps([d[1] for d in inventory_data])},
  marker: {{ colors: {json.dumps([d[2] for d in inventory_data])} }},
  textinfo: 'label+value+percent',
  textposition: 'auto',
  hovertemplate: '%{{label}}<br>%{{value}}개 (%{{percent}})<extra></extra>'
}}], {{
  showlegend: false, margin: {{t: 10, b: 10, l: 10, r: 10}},
  annotations: [{{ text: '{sum(cats.values())}<br>files', x: 0.5, y: 0.5,
                   font: {{size: 20, color: '#305496'}}, showarrow: false }}]
}}, {{responsive: true}});

// 2. 검증 진행률
Plotly.newPlot('chart_verify', [{{
  type: 'pie', hole: 0.55,
  labels: {json.dumps(list(verify_progress.keys()), ensure_ascii=False)},
  values: {json.dumps(list(verify_progress.values()))},
  marker: {{ colors: ['#5cb85c', '#e9ecef'] }},
  textinfo: 'label+value',
  hovertemplate: '%{{label}}<br>%{{value}}/%{{total}} (%{{percent}})<extra></extra>'
}}], {{
  showlegend: false, margin: {{t: 10, b: 10, l: 10, r: 10}},
  annotations: [{{ text: '{verified_count}/{total_estimate}', x: 0.5, y: 0.5,
                   font: {{size: 20, color: '#305496'}}, showarrow: false }}]
}}, {{responsive: true}});

// 3. 중복 그룹 분포
Plotly.newPlot('chart_dupsize', [{{
  type: 'bar',
  x: {json.dumps([f'{k}종' for k in sorted(dup_size_distribution.keys())])},
  y: {json.dumps([dup_size_distribution[k] for k in sorted(dup_size_distribution.keys())])},
  marker: {{ color: '#f0ad4e' }},
  text: {json.dumps([str(dup_size_distribution[k]) + '개 그룹' for k in sorted(dup_size_distribution.keys())], ensure_ascii=False)},
  textposition: 'outside',
  hovertemplate: '%{{x}} 동일: %{{y}}개 그룹<extra></extra>'
}}], {{
  xaxis: {{ title: '그룹 크기 (한 그룹 내 동일 이미지 수)' }},
  yaxis: {{ title: '그룹 개수' }},
  margin: {{t: 20, b: 60, l: 50, r: 20}}
}}, {{responsive: true}});

// 4. 펜타스톰 패밀리 트리 (sunburst)
Plotly.newPlot('chart_penta', [{{
  type: 'sunburst',
  labels: {json.dumps(penta_sunburst_data['labels'], ensure_ascii=False)},
  parents: {json.dumps(penta_sunburst_data['parents'], ensure_ascii=False)},
  values: {json.dumps(penta_sunburst_data['values'])},
  branchvalues: 'remainder',
  marker: {{ colors: ['#305496', '#5cb85c', '#5bc0de', '#a8d8a8', '#f0ad4e', '#d9534f', '#9bdaf3', '#8ad4eb', '#7ed6f1', '#5bc0de'] }},
  hovertemplate: '<b>%{{label}}</b><br>%{{value}} 변종<extra></extra>'
}}], {{
  margin: {{t: 10, b: 10, l: 10, r: 10}}
}}, {{responsive: true}});

// 5. Sankey
Plotly.newPlot('chart_sankey', [{{
  type: 'sankey',
  arrangement: 'snap',
  node: {{
    pad: 8, thickness: 14,
    line: {{ color: 'white', width: 0.5 }},
    label: {json.dumps(labels, ensure_ascii=False)},
    color: {json.dumps(node_colors)},
  }},
  link: {{
    source: {json.dumps(sankey_source)},
    target: {json.dumps(sankey_target)},
    value:  {json.dumps(sankey_value)},
    color: 'rgba(48, 84, 150, 0.25)',
  }}
}}], {{
  margin: {{t: 10, b: 10, l: 10, r: 10}},
  font: {{size: 11}}
}}, {{responsive: true}});

// 6. 카봇 스카이 변종-모드 매트릭스 (heatmap)
const skyData = [
  ['basic',           1, 1, 0, 0],
  ['swat',            1, 1, 0, 0],
  ['swat_x',          1, 1, 0, 0],
  ['swat_x_crystal',  1, 0, 0, 0],
];
Plotly.newPlot('chart_skymatrix', [{{
  type: 'heatmap',
  z: skyData.map(r => r.slice(1)),
  x: ['robot', 'vehicle', 'beast', 'power'],
  y: skyData.map(r => r[0]),
  colorscale: [[0, '#e9ecef'], [1, '#5cb85c']],
  showscale: false,
  text: skyData.map(r => r.slice(1).map(v => v ? '✓' : '—')),
  texttemplate: '%{{text}}',
  textfont: {{ size: 22, color: 'white' }},
  hovertemplate: '%{{y}} × %{{x}}: %{{z}}<extra></extra>',
  xgap: 4, ygap: 4
}}], {{
  xaxis: {{ side: 'top', title: '' }},
  yaxis: {{ title: '변종', autorange: 'reversed' }},
  margin: {{t: 50, b: 20, l: 120, r: 20}}
}}, {{responsive: true}});
</script>

</body>
</html>
"""
    out_path.write_text(html, encoding='utf-8')
    print(f"[html] {out_path}")


# ─────────────────────────────────────────────────────────────────────────
# 5. 메인
# ─────────────────────────────────────────────────────────────────────────

def main():
    rows, cats, dups = scan_inventory()
    print(f"인벤토리: {cats} / 중복 그룹 {len(dups)}개")

    build_excel(BASE / "carbot_catalog.xlsx")
    build_html(BASE / "carbot_visual.html", cats, dups)


if __name__ == "__main__":
    main()
